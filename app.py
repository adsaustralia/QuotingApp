
import streamlit as st
import pandas as pd
import numpy as np
import re, json, math, io
from pathlib import Path
from datetime import datetime
import openpyxl

APP_VERSION = "simple-ui-v4-openpyxl-stylecopy-safe"

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
MAPPING_DIR = DATA_DIR / "mappings"
MAPPING_DIR.mkdir(parents=True, exist_ok=True)
DEFAULT_STANDARD_STOCKS_CSV = DATA_DIR / "standard_stocks.csv"

def clean_text(s) -> str:
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("Ø", "dia ").replace("⌀", "dia ")
    return s

def load_mapping(customer: str) -> dict:
    fp = MAPPING_DIR / f"{clean_text(customer).replace(' ', '_')}_stock_map.json"
    if fp.exists():
        return json.loads(fp.read_text(encoding="utf-8"))
    return {"customer": customer, "mappings": {}}

def save_mapping(customer: str, mapping: dict) -> None:
    fp = MAPPING_DIR / f"{clean_text(customer).replace(' ', '_')}_stock_map.json"
    fp.write_text(json.dumps(mapping, indent=2), encoding="utf-8")

def parse_variant_spec(col_name: str):
    raw = "" if col_name is None else str(col_name)
    parts = [p.strip() for p in raw.split("\\n") if str(p).strip()]
    out = {"size_text": None, "colour": None, "sides_text": None, "stock_from_header": None, "finishing": None}
    if len(parts) >= 1: out["size_text"] = parts[0]
    if len(parts) >= 2: out["colour"] = parts[1]
    if len(parts) >= 3: out["sides_text"] = parts[2]
    if len(parts) >= 4: out["stock_from_header"] = parts[3]
    if len(parts) >= 5: out["finishing"] = parts[4]
    return out

def parse_size_text(text: str):
    t = clean_text(text)
    if not t:
        return {"shape":"unknown","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":np.nan}

    m = re.search(r"(dia(?:meter)?)\\s*[:=]?\\s*(\\d+(?:\\.\\d+)?)", t)
    if m:
        d = float(m.group(2))
        return {"shape":"circle","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":d}

    m = re.search(r"(\\d+(?:\\.\\d+)?)\\s*(x|\\*|by)\\s*(\\d+(?:\\.\\d+)?)", t)
    if m:
        w = float(m.group(1))
        h = float(m.group(3))
        return {"shape":"rectangle","width_mm":w,"height_mm":h,"diameter_mm":np.nan}

    return {"shape":"unknown","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":np.nan}

def sqm_calc(shape: str, width_mm=None, height_mm=None, diameter_mm=None):
    if shape == "rectangle" and pd.notna(width_mm) and pd.notna(height_mm):
        return (float(width_mm)/1000.0) * (float(height_mm)/1000.0)
    if shape == "circle" and pd.notna(diameter_mm):
        r = (float(diameter_mm)/1000.0)/2.0
        return math.pi * (r**2)
    return np.nan

def sides_normalize(val: str, default="SS"):
    t = clean_text(val)
    if t in ("ds","2s","double","2pp","two"):
        return "DS"
    if t in ("ss","1s","single","1pp","one"):
        return "SS"
    return default

def export_quote_excel(df_lines: pd.DataFrame, df_summary: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Quote Summary")
        df_lines.to_excel(writer, index=False, sheet_name="Line Items")
    return bio.getvalue()

def export_quote_pdf(df_summary: pd.DataFrame, df_lines: pd.DataFrame, title="Quote") -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm

    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    y = h - 18*mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(18*mm, y, title)
    y -= 7*mm
    c.setFont("Helvetica", 9)
    c.drawString(18*mm, y, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {APP_VERSION}")
    y -= 8*mm

    c.setFont("Helvetica-Bold", 11)
    c.drawString(18*mm, y, "Summary")
    y -= 6*mm
    c.setFont("Helvetica", 10)
    for _, row in df_summary.iterrows():
        c.drawString(20*mm, y, f"{row['Label']}: {row['Value']}")
        y -= 5*mm

    y -= 4*mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(18*mm, y, "Line Items (Material)")
    y -= 6*mm

    headers = ["Qty", "Sides", "Shape", "Size", "Stock", "Total SQM", "Rate", "Line Total"]
    col_x = [18, 32, 46, 62, 95, 150, 175, 195]  # mm
    c.setFont("Helvetica-Bold", 8)
    for hx, head in zip(col_x, headers):
        c.drawString(hx*mm, y, head)
    y -= 4*mm
    c.setFont("Helvetica", 8)

    def money(x):
        try: return f"{float(x):,.2f}"
        except: return str(x)

    for _, r in df_lines.head(80).iterrows():
        if y < 18*mm:
            c.showPage()
            y = h - 18*mm
            c.setFont("Helvetica-Bold", 8)
            for hx, head in zip(col_x, headers):
                c.drawString(hx*mm, y, head)
            y -= 4*mm
            c.setFont("Helvetica", 8)

        size_txt = r.get("size_display","")
        vals = [
            str(int(r.get("qty",0))),
            str(r.get("sides","")),
            str(r.get("shape","")),
            str(size_txt)[:16],
            str(r.get("stock_std",""))[:22],
            money(r.get("total_sqm",0)),
            money(r.get("sqm_rate",0)),
            money(r.get("line_total",0)),
        ]
        for hx, v in zip(col_x, vals):
            c.drawString(hx*mm, y, v)
        y -= 4*mm

    c.showPage()
    c.save()
    return bio.getvalue()

# ---------------- UI ----------------
st.set_page_config(page_title="Quoting App (Simple UI)", layout="wide")
st.title("Quoting App — Simple UI (W/H + Circle + Stock Mapping)")
st.caption("Pick only Size, Stock, Qty, Sides. Material = Total SQM × SQM Rate.")

with st.sidebar:
    st.header("Upload")
    customer = st.text_input("Customer", value="Adidas")
    uploaded = st.file_uploader("Customer Excel", type=["xlsx"])

    st.divider()
    st.header("Standard stock rates")
    st.caption("CSV columns required: stock_name_std, sqm_rate")
    std_upload = st.file_uploader("Optional: upload standard_stocks.csv", type=["csv"])

if uploaded is None:
    st.info("Upload an Excel file to start.")
    st.stop()

# Standard stocks
if std_upload is not None:
    df_std = pd.read_csv(std_upload)
else:
    if DEFAULT_STANDARD_STOCKS_CSV.exists():
        df_std = pd.read_csv(DEFAULT_STANDARD_STOCKS_CSV)
    else:
        df_std = pd.DataFrame(columns=["stock_name_std","sqm_rate"])

df_std["stock_name_std"] = df_std["stock_name_std"].astype(str)
df_std["sqm_rate"] = pd.to_numeric(df_std["sqm_rate"], errors="coerce")
std_options = df_std["stock_name_std"].dropna().tolist()
std_rate_map = dict(zip(df_std["stock_name_std"], df_std["sqm_rate"]))

# Sheets
wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
sheet_names = wb.sheetnames
wb.close()

top1, top2, top3 = st.columns([2,1,2])
sheet_name = top1.selectbox("Sheet", sheet_names, index=sheet_names.index("Standard_ WHS Pr MY AUS") if "Standard_ WHS Pr MY AUS" in sheet_names else 0)
header_row = top2.number_input("Header row (1-indexed)", 1, 100, 6, 1)
units = top3.selectbox("Units", ["mm","cm","m"], index=0)

u = {"mm":1.0,"cm":10.0,"m":1000.0}[units]
df_raw = pd.read_excel(uploaded, sheet_name=sheet_name, header=int(header_row)-1)

with st.expander("Preview (all rows)", expanded=False):
    st.dataframe(df_raw, use_container_width=True)

# ---------------- PICK VALUES (HORIZONTAL) ----------------
st.subheader("Pick Values")

c1, c2, c3, c4 = st.columns([2.2, 1.6, 2.0, 1.4])

with c1:
    st.markdown("**Size**")
    size_mode = st.radio("", ["W+H columns", "Size text column", "Circle diameter column"], index=0, horizontal=True, label_visibility="collapsed")
    if size_mode == "W+H columns":
        w_col = st.selectbox("Width column", df_raw.columns, index=df_raw.columns.get_loc("Width") if "Width" in df_raw.columns else 0)
        h_col = st.selectbox("Height column", df_raw.columns, index=df_raw.columns.get_loc("Height") if "Height" in df_raw.columns else 0)
        size_text_col = None
        dia_col = None
    elif size_mode == "Size text column":
        size_text_col = st.selectbox("Size column", df_raw.columns, index=df_raw.columns.get_loc("Description") if "Description" in df_raw.columns else 0)
        w_col = h_col = None
        dia_col = None
    else:
        dia_col = st.selectbox("Diameter column", df_raw.columns)
        w_col = h_col = None
        size_text_col = None

with c2:
    st.markdown("**Stock**")
    stock_col = st.selectbox("", df_raw.columns, index=df_raw.columns.get_loc("Stock") if "Stock" in df_raw.columns else 0, label_visibility="collapsed")

with c3:
    st.markdown("**Qty**")
    qty_mode = st.radio("", ["Single Qty column", "Multiple Qty columns (Adidas)"], index=1, horizontal=True, label_visibility="collapsed")
    if qty_mode == "Single Qty column":
        qty_col = st.selectbox("Qty column", df_raw.columns, index=df_raw.columns.get_loc("Qty") if "Qty" in df_raw.columns else 0)
        start_col = end_col = None
    else:
        start_col = st.selectbox("Start qty column", df_raw.columns)
        end_col = st.selectbox("End qty column", df_raw.columns, index=len(df_raw.columns)-1)
        qty_col = None

with c4:
    st.markdown("**Sides**")
    sides_mode = st.radio("", ["Default", "From column"], index=0, horizontal=True, label_visibility="collapsed")
    if sides_mode == "Default":
        sides_default = st.radio("Default sides", ["SS","DS"], index=0, horizontal=True)
        sides_col = None
    else:
        sides_col = st.selectbox("Sides column", df_raw.columns)
        sides_default = "SS"

with st.expander("Advanced (optional) — Adidas variant header parsing", expanded=False):
    prefer_header_size = st.checkbox("Prefer Size from qty column header when available", value=True)
    prefer_header_stock = st.checkbox("Prefer Stock from qty column header when available", value=True)
    prefer_header_sides = st.checkbox("Prefer Sides from qty column header when available", value=True)

# ---------------- NORMALIZE ----------------
st.subheader("Normalized Line Items")

base_cols_guess = [c for c in ["Line Reference","Description","Width","Height","Stock","Finishing","Colour","Sides","Orientation"] if c in df_raw.columns]
df = df_raw.copy()

if qty_mode == "Multiple Qty columns (Adidas)":
    cols = df.columns.tolist()
    i0 = cols.index(start_col)
    i1 = cols.index(end_col)
    if i0 > i1:
        i0, i1 = i1, i0
    qty_cols = cols[i0:i1+1]

    df = df.copy()
    df["origin_row"] = df.index

    lines = df.melt(
        id_vars=["origin_row"] + [c for c in base_cols_guess if c in df.columns],
        value_vars=qty_cols,
        var_name="variant_spec",
        value_name="qty"
    )
    lines["qty"] = pd.to_numeric(lines["qty"], errors="coerce").fillna(0)
    lines = lines[lines["qty"] > 0].copy()

    parsed = lines["variant_spec"].apply(parse_variant_spec).apply(pd.Series)
    lines = pd.concat([lines, parsed], axis=1)

    base_stock = lines[stock_col].astype(str) if stock_col in lines.columns else ""
    stock_from_header = lines["stock_from_header"].fillna("")
    lines["stock_customer"] = np.where((prefer_header_stock) & (stock_from_header.astype(str).str.len() > 0),
                                       stock_from_header.astype(str),
                                       base_stock)

    if sides_mode == "From column" and sides_col in lines.columns:
        sides_base = lines[sides_col].astype(str)
    else:
        sides_base = sides_default

    sides_from_header = lines["sides_text"].fillna("")
    sides_src = np.where((prefer_header_sides) & (sides_from_header.astype(str).str.len() > 0),
                         sides_from_header.astype(str),
                         sides_base)
    lines["sides"] = pd.Series(sides_src).apply(lambda x: sides_normalize(x, default=sides_default))

    if size_mode == "W+H columns" and w_col in lines.columns and h_col in lines.columns:
        lines["width_mm"] = pd.to_numeric(lines[w_col], errors="coerce") * u
        lines["height_mm"] = pd.to_numeric(lines[h_col], errors="coerce") * u
        lines["diameter_mm"] = np.nan
        lines["shape"] = "rectangle"
        lines["size_display"] = lines[w_col].astype(str) + " x " + lines[h_col].astype(str)
    elif size_mode == "Circle diameter column" and dia_col in lines.columns:
        lines["width_mm"] = np.nan
        lines["height_mm"] = np.nan
        lines["diameter_mm"] = pd.to_numeric(lines[dia_col], errors="coerce") * u
        lines["shape"] = "circle"
        lines["size_display"] = "DIA " + lines[dia_col].astype(str)
    else:
        txt = lines[size_text_col].astype(str) if (size_text_col in lines.columns) else lines.get("Description","").astype(str)
        geo = txt.apply(parse_size_text).apply(pd.Series)
        lines = pd.concat([lines, geo.rename(columns={"shape":"shape_parsed","width_mm":"w_parsed","height_mm":"h_parsed","diameter_mm":"d_parsed"})], axis=1)
        lines["width_mm"] = pd.to_numeric(lines["w_parsed"], errors="coerce") * u
        lines["height_mm"] = pd.to_numeric(lines["h_parsed"], errors="coerce") * u
        lines["diameter_mm"] = pd.to_numeric(lines["d_parsed"], errors="coerce") * u
        lines["shape"] = lines["shape_parsed"].replace({"unknown":"rectangle"})
        lines["size_display"] = txt

    if prefer_header_size:
        geo_h = lines["size_text"].apply(parse_size_text).apply(pd.Series)
        geo_h = geo_h.rename(columns={"shape":"shape_h","width_mm":"w_h","height_mm":"h_h","diameter_mm":"d_h"})
        lines = pd.concat([lines, geo_h], axis=1)

        has_rect = pd.to_numeric(lines["w_h"], errors="coerce").notna() & pd.to_numeric(lines["h_h"], errors="coerce").notna()
        has_circ = pd.to_numeric(lines["d_h"], errors="coerce").notna()

        lines.loc[has_rect, "shape"] = "rectangle"
        lines.loc[has_rect, "width_mm"] = pd.to_numeric(lines.loc[has_rect, "w_h"], errors="coerce") * u
        lines.loc[has_rect, "height_mm"] = pd.to_numeric(lines.loc[has_rect, "h_h"], errors="coerce") * u
        lines.loc[has_rect, "diameter_mm"] = np.nan
        lines.loc[has_rect, "size_display"] = lines.loc[has_rect, "size_text"].astype(str)

        lines.loc[has_circ, "shape"] = "circle"
        lines.loc[has_circ, "diameter_mm"] = pd.to_numeric(lines.loc[has_circ, "d_h"], errors="coerce") * u
        lines.loc[has_circ, "width_mm"] = np.nan
        lines.loc[has_circ, "height_mm"] = np.nan
        lines.loc[has_circ, "size_display"] = lines.loc[has_circ, "size_text"].astype(str)

else:
    lines = df.copy()
    lines["origin_row"] = lines.index
    lines["qty"] = pd.to_numeric(lines[qty_col], errors="coerce").fillna(0)
    lines = lines[lines["qty"] > 0].copy()
    lines["stock_customer"] = lines[stock_col].astype(str)
    lines["sides"] = sides_default if sides_mode == "Default" else lines[sides_col].astype(str).apply(lambda x: sides_normalize(x, default=sides_default))

    if size_mode == "W+H columns":
        lines["width_mm"] = pd.to_numeric(lines[w_col], errors="coerce") * u
        lines["height_mm"] = pd.to_numeric(lines[h_col], errors="coerce") * u
        lines["diameter_mm"] = np.nan
        lines["shape"] = "rectangle"
        lines["size_display"] = lines[w_col].astype(str) + " x " + lines[h_col].astype(str)
    elif size_mode == "Circle diameter column":
        lines["width_mm"] = np.nan
        lines["height_mm"] = np.nan
        lines["diameter_mm"] = pd.to_numeric(lines[dia_col], errors="coerce") * u
        lines["shape"] = "circle"
        lines["size_display"] = "DIA " + lines[dia_col].astype(str)
    else:
        txt = lines[size_text_col].astype(str)
        geo = txt.apply(parse_size_text).apply(pd.Series)
        lines["width_mm"] = pd.to_numeric(geo["width_mm"], errors="coerce") * u
        lines["height_mm"] = pd.to_numeric(geo["height_mm"], errors="coerce") * u
        lines["diameter_mm"] = pd.to_numeric(geo["diameter_mm"], errors="coerce") * u
        lines["shape"] = geo["shape"].replace({"unknown":"rectangle"})
        lines["size_display"] = txt

lines["sqm_each"] = lines.apply(lambda r: sqm_calc(r.get("shape","rectangle"), r.get("width_mm"), r.get("height_mm"), r.get("diameter_mm")), axis=1)
lines["total_sqm"] = pd.to_numeric(lines["sqm_each"], errors="coerce") * pd.to_numeric(lines["qty"], errors="coerce")

# ---------------- STOCK MAPPING ----------------
st.subheader("Stock Mapping (Customer name → Standard stock with sqm rate)")

mapping = load_mapping(customer)
unique_cs = sorted({s for s in lines["stock_customer"].dropna().astype(str).tolist() if clean_text(s)})

map_df = pd.DataFrame([{
    "customer_stock": cs,
    "standard_stock": mapping["mappings"].get(clean_text(cs), "")
} for cs in unique_cs])

edited = st.data_editor(
    map_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "standard_stock": st.column_config.SelectboxColumn("Standard stock", options=[""] + std_options)
    }
)

if st.button("Save mappings", type="primary"):
    new_map = load_mapping(customer)
    for _, r in edited.iterrows():
        cs_key = clean_text(r["customer_stock"])
        std = str(r["standard_stock"] or "").strip()
        if cs_key and std:
            new_map["mappings"][cs_key] = std
    save_mapping(customer, new_map)
    st.success("Mappings saved.")

mapping = load_mapping(customer)
lines["stock_std"] = lines["stock_customer"].apply(lambda x: mapping["mappings"].get(clean_text(x), ""))
lines["sqm_rate"] = lines["stock_std"].map(std_rate_map)
lines["line_total"] = pd.to_numeric(lines["total_sqm"], errors="coerce") * pd.to_numeric(lines["sqm_rate"], errors="coerce")

# ---------------- REVIEW + EXPORT ----------------
st.subheader("Quote Review")

missing = lines[(lines["stock_std"]=="") | (lines["sqm_rate"].isna())]
if len(missing) > 0:
    st.warning(f"{len(missing)} line(s) missing mapping/rate. Map above to calculate totals.")
else:
    st.success("All lines mapped.")

review = lines[["qty","sides","shape","size_display","stock_customer","stock_std","sqm_each","total_sqm","sqm_rate","line_total"]].copy()
for c in ["sqm_each","total_sqm","sqm_rate","line_total"]:
    review[c] = pd.to_numeric(review[c], errors="coerce")

st.dataframe(review.sort_values(["stock_std","shape"]).reset_index(drop=True), use_container_width=True)

total_sqm = float(review["total_sqm"].fillna(0).sum())
subtotal = float(review["line_total"].fillna(0).sum())

summary = pd.DataFrame([
    {"Label":"Customer", "Value": customer},
    {"Label":"Sheet", "Value": sheet_name},
    {"Label":"Total SQM", "Value": f"{total_sqm:,.3f}"},
    {"Label":"Subtotal (Material)", "Value": f"{subtotal:,.2f}"},
])

st.markdown("**Totals**")
st.table(summary)

# ---------------- EXPORT (Preserve customer format) ----------------
st.subheader("Export")

e1, e2, e3 = st.columns([2.2, 2.2, 1.6])
with e1:
    price_target = st.selectbox(
        "Where to put price in the exported Excel (column)",
        options=["Add new column at end"] + df_raw.columns.tolist()
    )
with e2:
    price_mode = st.radio(
        "Price output mode",
        ["Row total (sum of all qty variants)", "Per-qty-cell (price in each qty column)"],
        index=0,
        horizontal=True
    )
with e3:
    include_detail_sheets = st.checkbox("Include detail sheets (Line Items + Summary)", value=True)

def build_customer_format_export() -> pd.DataFrame:
    df_out = df_raw.copy()
    col_name = "Price" if price_target == "Add new column at end" else price_target

    # Ensure target column exists (create at end if needed)
    if col_name not in df_out.columns:
        df_out[col_name] = np.nan

    if price_mode.startswith("Row total"):
        # Sum line_total back to original row index
        # If melt path, 'lines' is melted and keeps original index via id_vars rows; we can use its current index alignment.
        # We'll attach an origin row id using merge on all base cols if possible; simplest: use line-level join by index from df_raw if present.
        # Since melt duplicates rows, we approximate by grouping using the base columns that exist in df_out.
        base_keys = [c for c in ["Line Reference", "Description", "Width", "Height", "Stock"] if c in df_out.columns and c in lines.columns]
        if len(base_keys) >= 1:
            row_totals = lines.groupby(base_keys)["line_total"].sum().reset_index()
            df_out = df_out.merge(row_totals, on=base_keys, how="left", suffixes=("", "_calc"))
            df_out[col_name] = df_out["line_total"].astype(float)
            df_out.drop(columns=["line_total"], inplace=True)
        else:
            # fallback: cannot find keys, set blank
            df_out[col_name] = np.nan

    else:
        # Per-qty-cell: only valid when using multiple qty columns
        if "variant_spec" in lines.columns:
            # Create a pivot table of prices aligned to the qty columns
            base_keys = [c for c in ["Line Reference", "Description", "Width", "Height", "Stock"] if c in df_out.columns and c in lines.columns]
            if len(base_keys) >= 1:
                pv = lines.pivot_table(index=base_keys, columns="variant_spec", values="line_total", aggfunc="sum")
                pv = pv.reset_index()
                df_out = df_out.merge(pv, on=base_keys, how="left", suffixes=("", ""))
                # Now, for each qty column that exists in df_out and in pivot columns, optionally leave it or overwrite?
                # We will NOT overwrite qty values. Instead, write prices into the chosen price column as JSON-like string if user selected a single column.
                if price_target != "Add new column at end" and price_target in df_raw.columns:
                    # Put row-wise total in selected column to avoid overwriting qty columns.
                    price_cols = [c for c in pv.columns if c not in base_keys]
                    df_out[col_name] = df_out[price_cols].sum(axis=1, numeric_only=True)
                    # drop pivot price cols to preserve format strictly
                    df_out.drop(columns=price_cols, inplace=True, errors="ignore")
                else:
                    # Add new column at end: create a compact string summary of per-cell prices
                    price_cols = [c for c in pv.columns if c not in base_keys]
                    def pack_row(r):
                        items = []
                        for pc in price_cols:
                            v = r.get(pc)
                            if pd.notna(v) and float(v) != 0:
                                items.append(f"{pc}: {float(v):.2f}")
                        return " | ".join(items)
                    df_out[col_name] = df_out.apply(pack_row, axis=1)
                    df_out.drop(columns=price_cols, inplace=True, errors="ignore")
            else:
                df_out[col_name] = np.nan
        else:
            df_out[col_name] = np.nan

    return df_out

def export_preserving_excel() -> bytes:
    """
    Export by editing the ORIGINAL workbook with openpyxl so the sheet structure/format stays the same.
    We only write price values into the chosen column.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # Build row totals by original row position
    if "variant_spec" in lines.columns:
        row_totals = lines.groupby("origin_row")["line_total"].sum()
    else:
        row_totals = lines.groupby("origin_row")["line_total"].sum()

    # Load original workbook from uploaded file (Streamlit uploader gives a file-like object)
    uploaded.seek(0)
    wb = load_workbook(uploaded)
    ws = wb[sheet_name]

    hdr_row = int(header_row)  # 1-indexed header row in Excel
    # Find target column index
    if price_target == "Add new column at end":
        target_col_idx = ws.max_column + 1
        header_cell = ws.cell(row=hdr_row, column=target_col_idx, value="Price")
        # copy style from previous header cell if possible (safe copy)
        if target_col_idx > 1:
            try:
                from copy import copy as _copy
                prev = ws.cell(row=hdr_row, column=target_col_idx-1)
                header_cell._style = _copy(prev._style)
                header_cell.font = _copy(prev.font)
                header_cell.fill = _copy(prev.fill)
                header_cell.border = _copy(prev.border)
                header_cell.alignment = _copy(prev.alignment)
                header_cell.number_format = prev.number_format
            except Exception:
                # If style copy fails, keep default style to avoid crashing
                pass
    else:
        # Locate existing column by matching the header cell value in header row
        target_col_idx = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=hdr_row, column=c).value
            if str(v).strip() == str(price_target).strip():
                target_col_idx = c
                break
        if target_col_idx is None:
            # fallback: add new at end if header not found
            target_col_idx = ws.max_column + 1
            ws.cell(row=hdr_row, column=target_col_idx, value="Price")

    # Write values into rows (data starts at hdr_row + 1)
    # Pandas df_raw has 0-based row index that maps to Excel row hdr_row+1+i
    for i, val in row_totals.items():
        excel_row = hdr_row + 1 + int(i)
        if val is None or (isinstance(val, float) and np.isnan(val)):
            continue
        cell = ws.cell(row=excel_row, column=target_col_idx)
        cell.value = float(val)

        # Optional: set number format to 2 decimals while keeping style
        cell.number_format = "0.00"

    # Optionally add detail sheets without altering the original sheet
    if include_detail_sheets:
        # Remove existing detail sheets if present to avoid duplicates
        for name in ["Quote Summary", "Line Items"]:
            if name in wb.sheetnames:
                del wb[name]
        ws_sum = wb.create_sheet("Quote Summary")
        for r_idx, row in enumerate(summary.itertuples(index=False), start=1):
            ws_sum.cell(row=r_idx, column=1, value=row.Label)
            ws_sum.cell(row=r_idx, column=2, value=row.Value)

        ws_li = wb.create_sheet("Line Items")
        for c_idx, col in enumerate(review.columns.tolist(), start=1):
            ws_li.cell(row=1, column=c_idx, value=col)
        for r_idx, row in enumerate(review.itertuples(index=False), start=2):
            for c_idx, v in enumerate(row, start=1):
                ws_li.cell(row=r_idx, column=c_idx, value=v)

    # Save workbook to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

b1, b2 = st.columns(2)
with b1:
    xbytes = export_preserving_excel()
    st.download_button(
        "Download Quote Excel (preserve format)",
        data=xbytes,
        file_name=f"Quote_{customer}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
with b2:
    pbytes = export_quote_pdf(summary, review, title=f"Quote - {customer}")
    st.download_button(
        "Download Quote PDF",
        data=pbytes,
        file_name=f"Quote_{customer}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mime="application/pdf"
    )

with st.expander("Version / Debug", expanded=False):
    st.write("APP VERSION:", APP_VERSION)
    st.write("Rows:", len(lines), " | Unique customer stocks:", len(unique_cs))
